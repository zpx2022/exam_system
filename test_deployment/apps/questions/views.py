"""

试题管理视图

"""

from rest_framework import viewsets, status

from rest_framework.decorators import action

from rest_framework.permissions import IsAuthenticated

from rest_framework.response import Response

from rest_framework.parsers import MultiPartParser, JSONParser

from utils.permissions import IsTeacher

from utils.exceptions import BusinessException

from .models import Question

from .serializers import QuestionSerializer, QuestionListSerializer

from django.conf import settings

import os

from django.core.files.storage import default_storage

from django.core.files.base import ContentFile

from django.db import transaction

import csv

import io

import json

import logging

from utils.gm_crypto import gm_crypto



logger = logging.getLogger(__name__)





class QuestionViewSet(viewsets.ModelViewSet):

    """试题视图集"""

    queryset = Question.objects.all()

    permission_classes = [IsAuthenticated, IsTeacher]

    

    def get_serializer_class(self):

        if self.action == 'list':

            return QuestionListSerializer

        return QuestionSerializer

    

    def get_queryset(self):

        """过滤查询集"""

        queryset = Question.objects.all()

        

        # 按课程过滤

        course_id = self.request.query_params.get('course')

        if course_id:

            queryset = queryset.filter(course_id=course_id)

        

        # 按题型过滤

        question_type = self.request.query_params.get('type')

        if question_type:

            queryset = queryset.filter(type=question_type)

        

        # 按创建者过滤

        if self.request.user.role == 1:  # 教师只能看自己的

            queryset = queryset.filter(created_by=self.request.user)

        

        return queryset

    

    def create(self, request, *args, **kwargs):

        """创建试题"""

        import logging

        logger = logging.getLogger(__name__)

        

        try:

            

            # 检查是否是加密数据格式（解密失败的情况）

            if isinstance(request.data, dict):

                if 'encrypted_key' in request.data and 'encrypted_data' in request.data:

                    logger.error("收到加密数据格式，说明解密失败！")

                    return Response({

                        'code': 400,

                        'message': '数据解密失败，请检查加密配置或联系管理员',

                        'data': None

                    }, status=status.HTTP_400_BAD_REQUEST)

            

            serializer = self.get_serializer(data=request.data)

            if not serializer.is_valid():

                # 记录验证错误

                logger.warning(f"验证失败: {serializer.errors}")

                return Response({

                    'code': 400,

                    'message': '; '.join([f"{k}: {', '.join(str(v) for v in (v if isinstance(v, list) else [v]))}" for k, v in serializer.errors.items()]),

                    'data': None

                }, status=status.HTTP_400_BAD_REQUEST)
            
            # 保存试题
            self.perform_create(serializer)

            return Response({

                'code': 200,

                'message': ' creating success',

                'data': {}

            }, status=status.HTTP_201_CREATED)
            
        except Exception as e:
            logger.error(f"创建试题失败: {str(e)}", exc_info=True)

            return Response({

                'code': 500,

                'message': f'创建失败: {str(e)}',

                'data': None

            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    

    def update(self, request, *args, **kwargs):

        """更新试题"""

        partial = kwargs.pop('partial', False)

        instance = self.get_object()

        # 对象级权限检查：只有创建者可以更新

        if instance.created_by != request.user:

            from rest_framework.exceptions import PermissionDenied

            raise PermissionDenied("无权修改此试题")

        serializer = self.get_serializer(instance, data=request.data, partial=partial)

        serializer.is_valid(raise_exception=True)

        self.perform_update(serializer)

        

        return Response({

            'code': 200,

            'message': '更新成功',

            'data': serializer.data

        })

    

    def destroy(self, request, *args, **kwargs):

        """删除试题"""

        instance = self.get_object()

        # 对象级权限检查：只有创建者可以删除

        if instance.created_by != request.user:

            from rest_framework.exceptions import PermissionDenied

            raise PermissionDenied("无权删除此试题")

        self.perform_destroy(instance)

        return Response({

            'code': 200,

            'message': '删除成功',

            'data': None

        })

    

    def list(self, request, *args, **kwargs):

        """试题列表"""

        queryset = self.filter_queryset(self.get_queryset())

        page = self.paginate_queryset(queryset)

        

        if page is not None:

            serializer = self.get_serializer(page, many=True)

            return self.get_paginated_response(serializer.data)

        

        serializer = self.get_serializer(queryset, many=True)

        return Response({

            'code': 200,

            'message': '获取成功',

            'data': serializer.data

        })

    

    def get_paginated_response(self, data):

        """自定义分页响应格式"""

        return Response({

            'code': 200,

            'message': '获取成功',

            'data': {

                'results': data,

                'count': self.paginator.page.paginator.count,

                'next': self.paginator.get_next_link(),

                'previous': self.paginator.get_previous_link(),

            }

        })

    

    def retrieve(self, request, *args, **kwargs):

        """试题详情"""

        instance = self.get_object()

        serializer = self.get_serializer(instance)

        return Response({

            'code': 200,

            'message': '获取成功',

            'data': serializer.data

        })

    

    @action(detail=False, methods=['post'])

    def upload_media(self, request):

        """上传媒体文件（视频/音频）"""

        if 'file' not in request.FILES:

            return Response({

                'code': 400,

                'message': '请选择文件',

                'data': None

            }, status=status.HTTP_400_BAD_REQUEST)

        

        file = request.FILES['file']

        

        # 验证文件类型

        allowed_types = ['video/', 'audio/', 'image/']

        if not any(file.content_type.startswith(t) for t in allowed_types):

            return Response({

                'code': 400,

                'message': '不支持的文件类型',

                'data': None

            }, status=status.HTTP_400_BAD_REQUEST)

        

        # 保存文件 - 直接保存到media目录，不使用questions子目录

        file_path = file.name  # 直接使用文件名，不加questions/前缀

        saved_path = default_storage.save(file_path, ContentFile(file.read()))

        

        # 返回文件名（已经不包含前缀）

        return Response({

            'code': 200,

            'message': '上传成功',

            'data': {

                'file_path': saved_path,  # 直接返回保存的文件名

                'url': f"{settings.MEDIA_URL}{saved_path}"

            }

        })



    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser, JSONParser])

    @transaction.atomic

    def bulk_import(self, request, *args, **kwargs):

        """批量导入试题（仅支持XLSX格式）"""

        

        rows_data = []

        

        # 检查是否是统一加密的题目数据（来自前端Excel/CSV解析+统一加密）

        if 'questions' in request.data and 'has_encryption' in request.data:

            try:

                questions = request.data.get('questions')

                if not isinstance(questions, list):

                    raise BusinessException('题目数据格式错误', code=400)

                

                # 数据已经在security_middleware中合并完成，直接使用
                rows_data = questions
                
                # 添加课程ID到每个题目
                course_id = request.data.get('course')
                if course_id:
                    for question in rows_data:
                        question['course'] = course_id

                        

            except Exception as e:

                logger.error(f"统一加密题目数据处理失败: {str(e)}")

                raise BusinessException(f'题目数据处理失败: {str(e)}', code=400)

                

        # 检查是否是JSON数据（来自前端Excel/CSV解析）

        elif 'data' in request.data:

            try:

                data_field = request.data.get('data')

                # JSONParser已经解析了JSON，所以data_field直接是Python对象

                # 如果是字符串，说明是文件上传方式，需要再次解析

                if isinstance(data_field, str):

                    rows_data = json.loads(data_field)

                else:

                    # 已经是Python对象（列表）

                    rows_data = data_field

                    

                if not isinstance(rows_data, list):

                    raise BusinessException('数据格式错误', code=400)

                    

                # 添加课程ID到每个题目

                course_id = request.data.get('course')

                if course_id:

                    for question in rows_data:

                        question['course'] = course_id

                        

            except (json.JSONDecodeError, TypeError) as e:

                raise BusinessException('JSON数据解析失败', code=400)

        

        # 检查是否是文件上传

        elif 'file' in request.data or len(request.FILES) > 0:

            # 优先从request.FILES获取文件

            if len(request.FILES) > 0:

                file_obj = request.FILES.get('file')

                if not file_obj:

                    # 尝试从request.data获取

                    file_obj = request.data.get('file')

            else:

                file_obj = request.data.get('file')

                

            if not file_obj:

                raise BusinessException('未上传文件', code=400)



            # 处理字符串和文件对象两种情况

            if hasattr(file_obj, 'name'):

                file_name = file_obj.name.lower()

            else:

                file_name = str(file_obj).lower()

            if not (file_name.endswith('.xlsx') or file_name.endswith('.xls')):

                raise BusinessException('只支持Excel文件格式（.xlsx/.xls）', code=400)



            # 处理Excel文件

            import openpyxl

            wb = openpyxl.load_workbook(file_obj)

            ws = wb.active

            

            # 获取表头

            header = [cell.value for cell in ws[1]]

            required_headers = ['type', 'content']

            if not all(h in header for h in required_headers):

                raise BusinessException(f"Excel文件头必须包含: {', '.join(required_headers)}", code=400)

            

            # 读取数据

            data_rows = []

            for row in ws.iter_rows(min_row=2, values_only=True):

                data_rows.append(row)

            

            # 转换为字典格式

            rows_data = [dict(zip(header, row)) for row in data_rows]

        else:

            raise BusinessException('请上传Excel文件或提供数据', code=400)



        results = {

            'success_count': 0,

            'failure_count': 0,

            'errors': []

        }

        

        for i, row_data in enumerate(rows_data, start=2):

            question_type = str(row_data.get('type', '')).strip()

            content = str(row_data.get('content', '')).strip()

            options_str = str(row_data.get('options', '')).strip()

            answer = str(row_data.get('answer', '')).strip()

            analysis = str(row_data.get('analysis', '')).strip()

            course_id = row_data.get('course') or request.data.get('course')



            # 基础验证

            if not question_type or not content:

                results['failure_count'] += 1

                results['errors'].append({'row': i, 'content': content[:50] + '...' if len(content) > 50 else content, 'message': '题型和题干不能为空'})

                continue

            

            # 验证课程

            course = None

            if course_id:

                try:

                    from apps.courses.models import Course

                    course = Course.objects.get(id=course_id)

                    # 验证教师是否有权限为该课程创建试题

                    if request.user.role == 1 and course.teacher != request.user:

                        results['failure_count'] += 1

                        results['errors'].append({'row': i, 'content': content[:50] + '...' if len(content) > 50 else content, 'message': f'您不是课程 "{course.name}" 的授课教师'})

                        continue

                except Course.DoesNotExist:

                    results['failure_count'] += 1

                    results['errors'].append({'row': i, 'content': content[:50] + '...' if len(content) > 50 else content, 'message': f'课程ID {course_id} 不存在'})

                    continue



            try:

                question_type = int(question_type)

                if question_type not in [1, 2, 3, 4]:

                    raise ValueError()

            except (ValueError, TypeError):

                results['failure_count'] += 1

                results['errors'].append({'row': i, 'content': content[:50] + '...' if len(content) > 50 else content, 'message': f'无效的题型: {question_type} (应为1-4)'})

                continue



            # 验证选项格式（单选/多选题需要）

            options = {}

            if question_type in [1, 2]:  # 单选/多选题需要选项

                if not options_str:

                    results['failure_count'] += 1

                    results['errors'].append({'row': i, 'content': content[:50] + '...' if len(content) > 50 else content, 'message': '单选/多选题必须包含选项'})

                    continue

                

                try:

                    options = json.loads(options_str.replace('""', '"'))

                    if not isinstance(options, dict):

                        raise ValueError()

                except (json.JSONDecodeError, ValueError):

                    results['failure_count'] += 1

                    results['errors'].append({'row': i, 'content': content[:50] + '...' if len(content) > 50 else content, 'message': '选项格式错误，应为JSON格式'})

                    continue



                # 验证答案

            if not answer:

                results['failure_count'] += 1

                results['errors'].append({'row': i, 'content': content[:50] + '...' if len(content) > 50 else content, 'message': '答案不能为空'})

                continue



                # 处理多选题答案格式

            processed_answer = answer

            if question_type == 2:  # 多选题

                # 将逗号分隔的答案转换为数组

                processed_answer = [ans.strip() for ans in answer.split(',') if ans.strip()]

                logger.info(f"[Bulk Import] Multi-choice answer processed: {answer} -> {processed_answer}")

                if not processed_answer:

                    results['failure_count'] += 1

                    results['errors'].append({'row': i, 'content': content[:50] + '...' if len(content) > 50 else content, 'message': '多选题答案不能为空'})

                    continue

            elif question_type == 3:  # 判断题

                # 将"正确"/"错误"转换为"A"/"B"

                if answer == '正确':

                    processed_answer = 'A'

                elif answer == '错误':

                    processed_answer = 'B'

                else:

                    results['failure_count'] += 1

                    results['errors'].append({'row': i, 'content': content[:50] + '...' if len(content) > 50 else content, 'message': '判断题答案必须是"正确"或"错误"'})

                    continue

                logger.info(f"[Bulk Import] Judgment answer processed: {answer} -> {processed_answer}")



            try:

                # 创建试题

                # 先确定options内容（明文形式）
                if question_type in [1, 2]:  # 单选/多选题
                    question_options = options  # 已经是解析后的字典
                elif question_type == 3:  # 判断题
                    question_options = {"A": "正确", "B": "错误"}
                else:  # 填空题等其他类型
                    question_options = {}

                # 创建题目对象（options先设为空字符串，避免明文存储）
                question = Question.objects.create(
                    type=question_type,
                    content=content,
                    options='',  # 先设为空字符串
                    answer='temp',  # 临时值，稍后设置加密答案
                    analysis=analysis or None,
                    course=course,
                    created_by=request.user
                )

                # 使用加密方法设置options（关键修复）
                question.set_options(question_options)
                
                # 设置加密后的答案
                question.set_answer(processed_answer)
                
                question.save()

                

                logger.info(f"[Bulk Import] Created question: {content[:50]}...")



                results['success_count'] += 1

            except Exception as e:

                results['failure_count'] += 1

                results['errors'].append({'row': i, 'content': content[:50] + '...' if len(content) > 50 else content, 'message': f'创建失败: {str(e)}'})



        return Response({

            'code': 200,

            'message': '导入完成',

            'data': results

        })



    @action(detail=False, methods=['delete'])

    @transaction.atomic

    def bulk_delete(self, request, *args, **kwargs):

        """

        批量删除试题

        """

        ids = request.data.get('ids', [])

        

        if not ids:

            return Response({'code': 400, 'message': '请提供要删除的试题ID列表'}, 

                           status=status.HTTP_400_BAD_REQUEST)

        

        if not isinstance(ids, list):

            return Response({'code': 400, 'message': 'ids参数必须是数组'}, 

                           status=status.HTTP_400_BAD_REQUEST)

        

        success_count = 0

        failed_count = 0

        failed_ids = []

        errors = []

        

        for question_id in ids:

            try:

                question = Question.objects.get(id=question_id)

                

                # 权限检查：教师只能删除自己创建的试题，管理员可以删除所有试题

                if request.user.role == 1 and question.created_by != request.user:

                    errors.append(f'试题ID {question_id}: 您只能删除自己创建的试题')

                    failed_count += 1

                    failed_ids.append(question_id)

                    continue

                

                # 检查关联数据

                from apps.papers.models import PaperQuestion

                

                # 检查试题是否被试卷使用

                paper_question_count = PaperQuestion.objects.filter(question=question).count()

                if paper_question_count > 0:

                    errors.append(f'试题ID {question_id}: 该试题被 {paper_question_count} 份试卷使用，无法删除')

                    failed_count += 1

                    failed_ids.append(question_id)

                    continue

                

                # 删除试题（包括媒体文件）

                if question.media_file:

                    try:

                        # 删除媒体文件

                        media_path = question.media_file.name if hasattr(question.media_file, 'name') else str(question.media_file)

                        if default_storage.exists(media_path):

                            default_storage.delete(media_path)

                            logger.info(f"[Bulk Delete] Deleted media file: {media_path}")

                    except Exception as e:

                        logger.warning(f"[Bulk Delete] Failed to delete media file for question {question_id}: {str(e)}")

                

                question.delete()

                success_count += 1

                logger.info(f"[Bulk Delete] Deleted question: {question_id}")

                

            except Question.DoesNotExist:

                errors.append(f'试题ID {question_id}: 试题不存在')

                failed_count += 1

                failed_ids.append(question_id)

            except Exception as e:

                errors.append(f'试题ID {question_id}: 删除失败 - {str(e)}')

                failed_count += 1

                failed_ids.append(question_id)

        

        return Response({

            'code': 200,

            'message': f'批量删除完成，成功 {success_count} 个，失败 {failed_count} 个',

            'data': {

                'success_count': success_count,

                'failed_count': failed_count,

                'failed_ids': failed_ids,

                'errors': errors

            }

        })









