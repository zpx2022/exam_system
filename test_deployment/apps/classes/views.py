"""
班级管理视图
"""
from rest_framework import viewsets
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import action
from rest_framework.parsers import MultiPartParser
from rest_framework.response import Response
from django.db import models, transaction
from django.db.models import Count
import csv
import io
import json

from apps.courses.models import Course
from utils.exceptions import BusinessException
from utils.permissions import IsAdmin, IsAdminOrTeacher
from utils.pagination import StandardResultsSetPagination
from .models import Class
from .serializers import ClassSerializer


class ClassViewSet(viewsets.ModelViewSet):
    """
    班级管理视图集 (管理员)
    提供了对班级的增、删、改、查功能，并处理与课程的关联。
    """
    # prefetch_related('courses') 用于优化性能，一次性获取所有关联的课程
    # annotate 用于优化学生数查询，避免N+1问题
    queryset = Class.objects.all().prefetch_related('courses').annotate(
        student_count=Count('students')
    ).order_by('-created_at')
    serializer_class = ClassSerializer
    permission_classes = [IsAuthenticated]  # 基础权限要求登录
    pagination_class = StandardResultsSetPagination

    def get_permissions(self):
        """
        根据操作类型设置不同权限：
        - GET操作：管理员和教师可访问
        - POST/PUT/DELETE操作：仅管理员可访问
        """
        if self.action in ['list', 'retrieve']:
            self.permission_classes = [IsAuthenticated, IsAdminOrTeacher]
        elif self.action in ['bulk_import', 'bulk_delete']:
            self.permission_classes = [IsAuthenticated, IsAdmin]
        else:
            self.permission_classes = [IsAuthenticated, IsAdmin]
        return super().get_permissions()

    def list(self, request, *args, **kwargs):
        """班级列表 - 返回标准响应格式"""
        queryset = self.filter_queryset(self.get_queryset())
        page = self.paginate_queryset(queryset)
        
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        
        serializer = self.get_serializer(queryset, many=True)
        return Response({
            'code': 200,
            'message': '获取成功',
            'data': serializer.data
        })
    
    def get_paginated_response(self, data):
        """自定义分页响应格式"""
        return Response({
            'code': 200,
            'message': '获取成功',
            'data': {
                'results': data,
                'count': self.paginator.page.paginator.count,
                'next': self.paginator.get_next_link(),
                'previous': self.paginator.get_previous_link(),
            }
        })

    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser])
    @transaction.atomic
    def bulk_import(self, request, *args, **kwargs):
        """批量导入班级（仅支持XLSX格式）"""
        
        rows_data = []
        
        # 检查是否是JSON数据（来自前端Excel解析）
        if 'data' in request.data:
            try:
                rows_data = json.loads(request.data.get('data'))
                if not isinstance(rows_data, list):
                    raise BusinessException('数据格式错误', code=400)
            except json.JSONDecodeError:
                raise BusinessException('JSON数据解析失败', code=400)
        
        # 检查是否是文件上传
        elif 'file' in request.data or len(request.FILES) > 0:
            # 优先从request.FILES获取文件
            if len(request.FILES) > 0:
                file_obj = request.FILES.get('file')
                if not file_obj:
                    # 尝试从request.data获取
                    file_obj = request.data.get('file')
            else:
                file_obj = request.data.get('file')
                
            if not file_obj:
                raise BusinessException('未上传文件', code=400)

            # 处理字符串和文件对象两种情况
            if hasattr(file_obj, 'name'):
                file_name = file_obj.name.lower()
            else:
                file_name = str(file_obj).lower()
            if not (file_name.endswith('.xlsx') or file_name.endswith('.xls')):
                raise BusinessException('只支持Excel文件格式（.xlsx/.xls）', code=400)

            # 处理Excel文件
            import openpyxl
            wb = openpyxl.load_workbook(file_obj)
            ws = wb.active
            
            # 获取表头
            header = [cell.value for cell in ws[1]]
            required_headers = ['name', 'major', 'institution']
            if not all(h in header for h in required_headers):
                raise BusinessException(f"Excel文件头必须包含: {', '.join(required_headers)}", code=400)
            
            # 读取数据
            data_rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                data_rows.append(row)
            
            # 转换为字典格式
            rows_data = [dict(zip(header, row)) for row in data_rows]
        else:
            raise BusinessException('请上传Excel文件或提供数据', code=400)

        results = {
            'success_count': 0,
            'failure_count': 0,
            'errors': []
        }

        course_cache = {c.code: c for c in Course.objects.all()}

        for i, row_data in enumerate(rows_data, start=2):
            name = (row_data.get('name') or '').strip()
            major = (row_data.get('major') or '').strip()
            institution = (row_data.get('institution') or '').strip()
            description = (row_data.get('description') or '').strip()
            course_codes_raw = (row_data.get('course_codes') or '').strip()

            if not name or not major or not institution:
                results['failure_count'] += 1
                results['errors'].append({'row': i, 'name': name, 'message': '班级名称、专业、所属机构不能为空'})
                continue

            if Class.objects.filter(name=name, major=major, institution=institution).exists():
                results['failure_count'] += 1
                results['errors'].append({'row': i, 'name': name, 'message': '班级已存在'})
                continue

            courses_to_set = []
            if course_codes_raw:
                codes = [c.strip() for c in course_codes_raw.split(',') if c.strip()]
                missing = [c for c in codes if c not in course_cache]
                if missing:
                    results['failure_count'] += 1
                    results['errors'].append({'row': i, 'name': name, 'message': f"课程代码不存在: {', '.join(missing)}"})
                    continue
                courses_to_set = [course_cache[c] for c in codes]

            try:
                cls = Class.objects.create(
                    name=name,
                    major=major,
                    institution=institution,
                    description=description or None
                )
                if courses_to_set:
                    cls.courses.set(courses_to_set)
                results['success_count'] += 1
            except Exception as e:
                results['failure_count'] += 1
                results['errors'].append({'row': i, 'name': name, 'message': f'创建失败: {str(e)}'})

        return Response({
            'code': 200,
            'message': '导入完成',
            'data': results
        })

    @action(detail=False, methods=['delete'])
    @transaction.atomic
    def bulk_delete(self, request, *args, **kwargs):
        """
        批量删除班级
        """
        ids = request.data.get('ids', [])
        
        if not ids:
            return Response({'code': 400, 'message': '请提供要删除的班级ID列表'}, 
                           status=400)
        
        if not isinstance(ids, list):
            return Response({'code': 400, 'message': 'ids参数必须是数组'}, 
                           status=400)
        
        # 只有管理员可以批量删除
        if not request.user.is_staff and request.user.role != 0:
            return Response({'code': 403, 'message': '只有管理员可以批量删除班级'}, 
                           status=403)
        
        success_count = 0
        failed_count = 0
        failed_ids = []
        errors = []
        
        for class_id in ids:
            try:
                cls = Class.objects.get(id=class_id)
                
                # 检查关联数据
                from apps.users.models import User
                from apps.papers.models import Paper
                
                # 检查是否有学生属于该班级
                student_count = cls.students.count()
                if student_count > 0:
                    errors.append(f'班级ID {class_id}: 该班级有 {student_count} 个学生，无法删除')
                    failed_count += 1
                    failed_ids.append(class_id)
                    continue
                
                # 检查是否有试卷指定给该班级
                paper_count = Paper.objects.filter(target_classes=cls).count()
                if paper_count > 0:
                    errors.append(f'班级ID {class_id}: 该班级被 {paper_count} 份试卷指定，无法删除')
                    failed_count += 1
                    failed_ids.append(class_id)
                    continue
                
                # 删除班级
                cls.delete()
                success_count += 1
                
            except Class.DoesNotExist:
                errors.append(f'班级ID {class_id}: 班级不存在')
                failed_count += 1
                failed_ids.append(class_id)
            except Exception as e:
                errors.append(f'班级ID {class_id}: 删除失败 - {str(e)}')
                failed_count += 1
                failed_ids.append(class_id)
        
        return Response({
            'code': 200,
            'message': f'批量删除完成，成功 {success_count} 个，失败 {failed_count} 个',
            'data': {
                'success_count': success_count,
                'failed_count': failed_count,
                'failed_ids': failed_ids,
                'errors': errors
            }
        })

    def get_queryset(self):
        """
        重写查询集，以支持按名称、专业、学院进行搜索和筛选。
        """
        queryset = super().get_queryset()
        
        # 模糊搜索
        search = self.request.query_params.get('search')
        if search:
            queryset = queryset.filter(
                models.Q(name__icontains=search) |
                models.Q(major__icontains=search) |
                models.Q(institution__icontains=search)
            )
        
        # 精确筛选
        major = self.request.query_params.get('major')
        if major:
            queryset = queryset.filter(major__exact=major)
        
        institution = self.request.query_params.get('institution')
        if institution:
            queryset = queryset.filter(institution__exact=institution)
        
        return queryset

    def update(self, request, *args, **kwargs):
        """重写update方法，确保部分更新时使用partial=True"""
        partial = kwargs.pop('partial', True)  # 默认使用部分更新
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        
        return Response({
            'code': 200,
            'message': '更新成功',
            'data': serializer.data
        })