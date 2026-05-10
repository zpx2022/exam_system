import json
import logging

from django.contrib.auth import authenticate
from django.views.decorators.csrf import csrf_exempt
from django.conf import settings
from rest_framework import status, viewsets
from rest_framework.decorators import api_view, permission_classes, action
from rest_framework.parsers import MultiPartParser, JSONParser
from apps.classes.models import Class
from drf_spectacular.utils import extend_schema, inline_serializer
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework.authtoken.models import Token

from .models import User
from .serializers import (LoginSerializer, RegisterSerializer, UserInfoSerializer, 
                          AdminUserSerializer, ChangePasswordSerializer,
                          ForgotPasswordSendCodeSerializer, ForgotPasswordResetSerializer)
from utils.gm_crypto import gm_crypto
from utils.exceptions import BusinessException
from utils.permissions import IsAdmin
from utils.pagination import StandardResultsSetPagination
from django.db import transaction
import csv
import io
import string
from utils.redis_client import redis_client
import random

logger = logging.getLogger(__name__)


@extend_schema(
    request=RegisterSerializer,
    responses={
        201: inline_serializer(
            name='RegisterResponse',
            fields={
                'code': __import__('rest_framework').serializers.IntegerField(),
                'message': __import__('rest_framework').serializers.CharField(),
                'data': inline_serializer(
                    name='RegisterResponseData',
                    fields={
                        'id': __import__('rest_framework').serializers.IntegerField(),
                        'username': __import__('rest_framework').serializers.CharField(),
                    },
                ),
            },
        )
    },
)
@api_view(['POST'])
@permission_classes([AllowAny])
def register(request):
    """
    学生注册（统一分层解密）
    """
    serializer = RegisterSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)

    # 中间件已自动解密，直接使用解密后的数据
    username = serializer.validated_data.get('username')
    password = serializer.validated_data.get('password')
    real_name_plain = serializer.validated_data.get('real_name_plain', '')
    phone_plain = serializer.validated_data.get('phone_plain', '')

    if User.objects.filter(username=username).exists():
        raise BusinessException("用户名已存在", code=400)

    user = User.objects.create_user(username, password=password, role=2)
    if real_name_plain:
        user.set_real_name(real_name_plain)
    if phone_plain:
        user.set_phone(phone_plain)
    user.save()

    return Response({'code': 200, 'message': '注册成功', 'data': {'id': user.id, 'username': user.username}}, status=status.HTTP_201_CREATED)


@extend_schema(
    request=LoginSerializer,
    responses={
        200: inline_serializer(
            name='LoginResponse',
            fields={
                'code': __import__('rest_framework').serializers.IntegerField(),
                'message': __import__('rest_framework').serializers.CharField(),
                'data': inline_serializer(
                    name='LoginResponseData',
                    fields={
                        'token': __import__('rest_framework').serializers.CharField(),
                        'user': UserInfoSerializer(),
                    }
                )
            }
        )
    }
)
@api_view(['POST'])
@permission_classes([AllowAny])
def login(request):
    """
    用户登录（统一分层解密）
    """
    serializer = LoginSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)

    # 中间件已自动解密，直接使用解密后的数据
    username = serializer.validated_data.get('username')
    password = serializer.validated_data.get('password')

    # 验证用户名和密码
    user = authenticate(username=username, password=password)
    if user:
        token, _ = Token.objects.get_or_create(user=user)
        user_info = UserInfoSerializer(user).data
        return Response({
            'code': 200,
            'message': '登录成功',
            'data': {
                'token': token.key,
                'user': user_info
            }
        })
    else:
        raise BusinessException("用户名或密码错误", code=400)


@extend_schema(
    request=None,
    responses={
        200: inline_serializer(
            name='LogoutResponse',
            fields={
                'code': __import__('rest_framework').serializers.IntegerField(),
                'message': __import__('rest_framework').serializers.CharField(),
            },
        )
    },
)
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def logout(request):
    """
    用户登出
    """
    request.user.auth_token.delete()
    return Response({'code': 200, 'message': '登出成功'}, status=status.HTTP_200_OK)


@extend_schema(
    request=None,
    responses={
        200: inline_serializer(
            name='UserInfoResponse',
            fields={
                'code': __import__('rest_framework').serializers.IntegerField(),
                'message': __import__('rest_framework').serializers.CharField(),
                'data': UserInfoSerializer(),
            },
        )
    },
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def user_info(request):
    """
    获取当前用户信息
    """
    serializer = UserInfoSerializer(request.user)
    return Response({'code': 200, 'message': '获取成功', 'data': serializer.data})


@extend_schema(
    request=inline_serializer(
        name='UpdateUserInfoRequest',
        fields={
            'real_name': __import__('rest_framework').serializers.CharField(required=False, allow_blank=True),
            'phone': __import__('rest_framework').serializers.CharField(required=False, allow_blank=True),
        },
    ),
    responses={
        200: inline_serializer(
            name='UpdateUserInfoResponse',
            fields={
                'code': __import__('rest_framework').serializers.IntegerField(),
                'message': __import__('rest_framework').serializers.CharField(),
                'data': UserInfoSerializer(),
            },
        )
    },
)
@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def update_user_info(request):
    """
    更新用户信息 (只能更新real_name和phone)
    """
    user = request.user
    real_name = request.data.get('real_name')
    phone = request.data.get('phone')
    if real_name is not None:
        user.set_real_name(real_name)
    if phone is not None:
        user.set_phone(phone)
    user.save()
    serializer = UserInfoSerializer(user)
    return Response({'code': 200, 'message': '更新成功', 'data': serializer.data})


@extend_schema(
    request=ChangePasswordSerializer,
    responses={
        200: inline_serializer(
            name='ChangePasswordResponse',
            fields={
                'code': __import__('rest_framework').serializers.IntegerField(),
                'message': __import__('rest_framework').serializers.CharField(),
            },
        )
    },
)
@csrf_exempt
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def send_password_change_code(request):
    """发送修改密码验证码（开发环境写日志）"""
    user = request.user

    phone = user.get_phone()
    if not phone:
        raise BusinessException('未绑定手机号，无法发送验证码', code=400)

    if redis_client.check_password_change_cooldown(user.id):
        raise BusinessException('操作过于频繁，请稍后再试', code=429)

    code = ''.join(str(random.randint(0, 9)) for _ in range(6))
    redis_client.set_password_change_code(user.id, code, ttl=300)
    redis_client.set_password_change_cooldown(user.id, ttl=60)

    if settings.DEBUG:
        logger.info('[修改密码验证码] user_id=%s phone=%s code=%s', user.id, phone, code)

    return Response({'code': 200, 'message': '验证码已发送'})


@csrf_exempt
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def change_password(request):
    """
    修改密码（旧密码 + 手机验证码）
    """
    serializer = ChangePasswordSerializer(data=request.data, context={'request': request})
    serializer.is_valid(raise_exception=True)

    cached_code = redis_client.get_password_change_code(request.user.id)
    if not cached_code:
        raise BusinessException('验证码已过期或不存在', code=400)
    if str(serializer.validated_data['code']).strip() != str(cached_code).strip():
        raise BusinessException('验证码错误', code=400)

    request.user.set_password(serializer.validated_data['new_password'])
    request.user.save()
    redis_client.delete_password_change_code(request.user.id)

    return Response({'code': 200, 'message': '密码修改成功'})


@extend_schema(
    request=None,
    responses={
        200: inline_serializer(
            name='AdminDashboardStatsResponse',
            fields={
                'code': __import__('rest_framework').serializers.IntegerField(),
                'message': __import__('rest_framework').serializers.CharField(),
                'data': inline_serializer(
                    name='AdminDashboardStatsData',
                    fields={
                        'user_count': __import__('rest_framework').serializers.IntegerField(),
                        'teacher_count': __import__('rest_framework').serializers.IntegerField(),
                        'student_count': __import__('rest_framework').serializers.IntegerField(),
                    },
                ),
            },
        )
    },
)
@api_view(['GET'])
@permission_classes([IsAuthenticated, IsAdmin])
def admin_dashboard_stats(request):
    """
    管理员仪表盘统计数据
    """
    stats = {
        'user_count': User.objects.count(),
        'teacher_count': User.objects.filter(role=1).count(),
        'student_count': User.objects.filter(role=2).count(),
    }
    return Response({'code': 200, 'message': '获取成功', 'data': stats})


class AdminUserViewSet(viewsets.ModelViewSet):
    """
    管理员用户管理接口
    """
    queryset = User.objects.all().order_by('-date_joined')
    serializer_class = AdminUserSerializer
    permission_classes = [IsAuthenticated, IsAdmin]
    pagination_class = StandardResultsSetPagination

    def create(self, request, *args, **kwargs):
        """创建用户"""
        return super().create(request, *args, **kwargs)

    @extend_schema(parameters=__import__('apps.users.schema', fromlist=['admin_user_list_extra_params']).admin_user_list_extra_params)
    def list(self, request, *args, **kwargs):
        """用户列表 - 返回标准响应格式"""
        queryset = self.filter_queryset(self.get_queryset())
        page = self.paginate_queryset(queryset)
        
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        
        serializer = self.get_serializer(queryset, many=True)
        return Response({
            'code': 200,
            'message': '获取成功',
            'data': serializer.data
        })
    
    def get_paginated_response(self, data):
        """自定义分页响应格式"""
        return Response({
            'code': 200,
            'message': '获取成功',
            'data': {
                'results': data,
                'count': self.paginator.page.paginator.count,
                'next': self.paginator.get_next_link(),
                'previous': self.paginator.get_previous_link(),
            }
        })
    
    def get_queryset(self):
        """支持搜索和过滤"""
        queryset = super().get_queryset()
        search = self.request.query_params.get('search')
        role = self.request.query_params.get('role')
        is_active = self.request.query_params.get('is_active')
        
        if search:
            queryset = queryset.filter(username__icontains=search)
        if role is not None:
            try:
                role = int(role)
                queryset = queryset.filter(role=role)
            except (ValueError, TypeError):
                pass
        if is_active is not None:
            is_active_bool = str(is_active).lower() in ('true', '1', 'yes')
            queryset = queryset.filter(is_active=is_active_bool)
        
        return queryset

    @action(detail=False, methods=['post'], parser_classes=[JSONParser])
    @transaction.atomic
    def bulk_import(self, request, *args, **kwargs):
        """批量导入用户（支持JSON格式）"""
        
        rows_data = []
        
        # 检查是否是统一加密的用户数据（来自前端Excel/CSV解析+统一加密）
        if 'users' in request.data and 'has_encryption' in request.data:
            try:
                users = request.data.get('users')
                if not isinstance(users, list):
                    raise BusinessException('用户数据格式错误', code=400)
                
                # 数据已经在security_middleware中合并完成，直接使用
                rows_data = users
                        
            except Exception as e:
                logger.error(f"统一加密用户数据处理失败: {str(e)}")
                raise BusinessException(f'用户数据处理失败: {str(e)}', code=400)
                
        # 检查是否是加密用户数据（来自前端Excel/CSV解析+加密）- 旧版本兼容
        elif 'encrypted_users' in request.data:
            try:
                encrypted_users = request.data.get('encrypted_users')
                if not isinstance(encrypted_users, list):
                    raise BusinessException('用户数据格式错误', code=400)
                
                # 解密每个用户的数据
                rows_data = []
                for i, encrypted_user in enumerate(encrypted_users):
                    try:
                        # 检查这个用户是否有加密数据
                        if 'encrypted_key' in encrypted_user and 'encrypted_data' in encrypted_user:
                            # 解密这个用户的数据
                            encrypted_key_hex = encrypted_user['encrypted_key']
                            encrypted_data_hex = encrypted_user['encrypted_data']
                            
                            # SM2解密密钥
                            sm4_key = gm_crypto.sm2_decrypt_key(encrypted_key_hex)
                            # SM4解密数据
                            decrypted_data_str = gm_crypto.sm4_decrypt_data(encrypted_data_hex, sm4_key)
                            decrypted_data = json.loads(decrypted_data_str)
                            
                            # 合并非敏感数据和解密的敏感数据
                            user_data = encrypted_user.copy()
                            user_data.pop('encrypted_key', None)
                            user_data.pop('encrypted_data', None)
                            user_data.update(decrypted_data)
                            
                            rows_data.append(user_data)
                        else:
                            # 这个用户没有加密数据，直接使用
                            rows_data.append(encrypted_user)
                            
                    except Exception as e:
                        logger.error(f"解密第{i+1}个用户数据失败: {str(e)}")
                        # 添加错误信息到结果中
                        rows_data.append({
                            'username': encrypted_user.get('username', f'unknown_{i+1}'),
                            'error': f'数据解密失败: {str(e)}'
                        })
                        
            except Exception as e:
                raise BusinessException(f'加密用户数据处理失败: {str(e)}', code=400)
                
        # 检查是否是JSON数据（来自前端Excel/CSV解析）
        elif 'data' in request.data:
            try:
                data_field = request.data.get('data')
                # JSONParser已经解析了JSON，所以data_field直接是Python对象
                # 如果是字符串，说明是文件上传方式，需要再次解析
                if isinstance(data_field, str):
                    rows_data = json.loads(data_field)
                else:
                    # 已经是Python对象（列表）
                    rows_data = data_field
                    
                if not isinstance(rows_data, list):
                    raise BusinessException('数据格式错误', code=400)
            except (json.JSONDecodeError, TypeError) as e:
                raise BusinessException('JSON数据解析失败', code=400)
        
        # 检查是否是文件上传
        elif 'file' in request.data or len(request.FILES) > 0:
            # 优先从request.FILES获取文件
            if len(request.FILES) > 0:
                file_obj = request.FILES.get('file')
                if not file_obj:
                    # 尝试从request.data获取
                    file_obj = request.data.get('file')
            else:
                file_obj = request.data.get('file')
                
            if not file_obj:
                raise BusinessException('未上传文件', code=400)

            # 处理字符串和文件对象两种情况
            if hasattr(file_obj, 'name'):
                file_name = file_obj.name.lower()
            else:
                file_name = str(file_obj).lower()
            
            if file_name.endswith('.xlsx') or file_name.endswith('.xls'):
                # 处理Excel文件
                import openpyxl
                wb = openpyxl.load_workbook(file_obj)
                ws = wb.active
                
                # 获取表头
                header = [cell.value for cell in ws[1]]
                required_headers = ['username', 'real_name', 'role']
                if not all(h in header for h in required_headers):
                    raise BusinessException(f"Excel文件头必须包含: {', '.join(required_headers)}", code=400)
                
                # 读取数据
                data_rows = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    data_rows.append(row)
                
                # 转换为字典格式
                rows_data = [dict(zip(header, row)) for row in data_rows]
                
            elif file_name.endswith('.csv'):
                # 处理CSV文件
                decoded_file = file_obj.read().decode('utf-8')
                io_string = io.StringIO(decoded_file)
                reader = csv.reader(io_string)
                
                header = next(reader)
                required_headers = ['username', 'real_name', 'role']
                if not all(h in header for h in required_headers):
                    raise BusinessException(f"CSV文件头必须包含: {', '.join(required_headers)}", code=400)

                rows_data = []
                for row in reader:
                    rows_data.append(dict(zip(header, row)))
            else:
                raise BusinessException('只支持Excel文件（.xlsx/.xls）和CSV文件格式', code=400)
        else:
            # 既没有data也没有file
            raise BusinessException('请提供Excel文件或JSON数据', code=400)
        
        results = {
            'success_count': 0,
            'failure_count': 0,
            'errors': [],
            'imported_users': {
                'teachers': [],  # 教师（role=1）
                'students': []   # 学生（role=2）
            }
        }
        
        class_cache = {c.name: c for c in Class.objects.all()}

        for i, row_data in enumerate(rows_data, start=1):
            username = str(row_data.get('username', '')).strip()
            real_name = str(row_data.get('real_name', '')).strip()
            role_str = str(row_data.get('role', '')).strip()
            phone = str(row_data.get('phone', '')).strip()  # 确保phone是字符串
            class_name = str(row_data.get('student_class_name', '')).strip()  # 确保class_name是字符串

            if not username or not real_name or not role_str:
                results['failure_count'] += 1
                results['errors'].append({'row': i, 'username': username, 'message': '用户名、真实姓名和角色不能为空'})
                continue

            if User.objects.filter(username=username).exists():
                results['failure_count'] += 1
                results['errors'].append({'row': i, 'username': username, 'message': '用户名已存在'})
                continue
            
            try:
                role = int(role_str)
                if role not in [1, 2]:
                     raise ValueError()
            except (ValueError, TypeError):
                results['failure_count'] += 1
                results['errors'].append({'row': i, 'username': username, 'message': f'无效的角色值: {role_str} (应为1或2)'})
                continue

            student_class = None
            if role == 2 and class_name:
                student_class = class_cache.get(class_name)
                if not student_class:
                    results['failure_count'] += 1
                    results['errors'].append({'row': i, 'username': username, 'message': f'班级 "{class_name}" 不存在'})
                    continue
            
            try:
                # 生成随机密码（使用更兼容的方法）
                password = ''.join(random.choice(string.ascii_letters + string.digits) for _ in range(8))
                
                user = User.objects.create_user(username=username, password=password, role=role)
                user.set_real_name(real_name)
                if phone:
                    user.set_phone(phone)
                if student_class:
                    user.student_class = student_class
                user.save()
                

                # 根据角色分别存储用户信息
                user_info = {
                    'username': username,
                    'password': password,
                    'role': role,
                    'real_name': real_name,
                    'phone': phone or '',
                    'student_class_name': class_name or ''
                }
                
                if role == 1:
                    results['imported_users']['teachers'].append(user_info)
                elif role == 2:
                    results['imported_users']['students'].append(user_info)
                results['success_count'] += 1
            except Exception as e:
                logger.error(f"[Bulk Import] Error creating user '{username}': {str(e)}")
                results['failure_count'] += 1
                results['errors'].append({'row': i, 'username': username, 'message': f'创建失败: {str(e)}'})

        return Response({
            'code': 200,
            'message': '导入完成',
            'data': results
        })

    @action(detail=False, methods=['delete'])
    @transaction.atomic
    def bulk_delete(self, request, *args, **kwargs):
        """
        批量删除用户
        """
        ids = request.data.get('ids', [])
        
        if not ids:
            return Response({'code': 400, 'message': '请提供要删除的用户ID列表'}, 
                           status=status.HTTP_400_BAD_REQUEST)
        
        if not isinstance(ids, list):
            return Response({'code': 400, 'message': 'ids参数必须是数组'}, 
                           status=status.HTTP_400_BAD_REQUEST)
        
        success_count = 0
        failed_count = 0
        failed_ids = []
        errors = []
        
        for user_id in ids:
            try:
                user = User.objects.get(id=user_id)
                
                # 不能删除自己
                if user.id == request.user.id:
                    errors.append(f'用户ID {user_id}: 不能删除自己')
                    failed_count += 1
                    failed_ids.append(user_id)
                    continue
                
                # 检查关联数据
                from apps.exams.models import ExamRecord
                from apps.papers.models import Paper
                
                # 检查是否有考试记录
                exam_count = ExamRecord.objects.filter(student=user).count()
                if exam_count > 0:
                    errors.append(f'用户ID {user_id}: 该用户有 {exam_count} 条考试记录，无法删除')
                    failed_count += 1
                    failed_ids.append(user_id)
                    continue
                
                # 检查是否创建了试卷
                paper_count = Paper.objects.filter(created_by=user).count()
                if paper_count > 0:
                    errors.append(f'用户ID {user_id}: 该用户创建了 {paper_count} 份试卷，无法删除')
                    failed_count += 1
                    failed_ids.append(user_id)
                    continue
                
                # 删除用户
                user.delete()
                success_count += 1
                
            except User.DoesNotExist:
                errors.append(f'用户ID {user_id}: 用户不存在')
                failed_count += 1
                failed_ids.append(user_id)
            except Exception as e:
                errors.append(f'用户ID {user_id}: 删除失败 - {str(e)}')
                failed_count += 1
                failed_ids.append(user_id)
        
        return Response({
            'code': 200,
            'message': f'批量删除完成，成功 {success_count} 个，失败 {failed_count} 个',
            'data': {
                'success_count': success_count,
                'failed_count': failed_count,
                'failed_ids': failed_ids,
                'errors': errors
            }
        })


@extend_schema(
    request=None,
    responses={
        200: inline_serializer(
            name='GetPublicKeyResponse',
            fields={
                'code': __import__('rest_framework').serializers.IntegerField(),
                'message': __import__('rest_framework').serializers.CharField(),
                'data': inline_serializer(
                    name='GetPublicKeyData',
                    fields={
                        'public_key': __import__('rest_framework').serializers.CharField(allow_null=True),
                    },
                ),
            },
        )
    },
)
@api_view(['GET'])
@permission_classes([AllowAny])
def get_public_key(request):
    """
    提供 SM2 公钥
    """
    return Response({
        'code': 200,
        'message': '获取成功',
        'data': {
            'public_key': settings.SM2_PUBLIC_KEY
        }
    })


@csrf_exempt
@api_view(['POST'])
@permission_classes([AllowAny])
def send_forgot_password_code(request):
    """发送忘记密码验证码"""
    serializer = ForgotPasswordSendCodeSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    username = serializer.validated_data['username']

    try:
        user = User.objects.get(username=username)
    except User.DoesNotExist:
        # 为防止用户枚举，即使不存在也返回成功
        logger.warning(f'Attempt to send forgot password code for non-existent user: {username}')
        return Response({'code': 200, 'message': '验证码已发送'})

    phone = user.get_phone()
    if not phone:
        raise BusinessException('该用户未绑定手机号，请联系管理员', code=400)

    if redis_client.check_forgot_password_cooldown(user.id):
        raise BusinessException('操作过于频繁，请稍后再试', code=429)

    code = ''.join(str(random.randint(0, 9)) for _ in range(6))
    redis_client.set_forgot_password_code(user.id, code, ttl=300)
    redis_client.set_forgot_password_cooldown(user.id, ttl=60)

    if settings.DEBUG:
        logger.info('[忘记密码验证码] user_id=%s phone=%s code=%s', user.id, phone, code)

    return Response({'code': 200, 'message': '验证码已发送'})


@csrf_exempt
@api_view(['POST'])
@permission_classes([AllowAny])
def reset_password(request):
    """重置密码"""
    serializer = ForgotPasswordResetSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)

    username = serializer.validated_data['username']
    code = serializer.validated_data['code']
    new_password = serializer.validated_data['new_password']

    try:
        user = User.objects.get(username=username)
    except User.DoesNotExist:
        raise BusinessException('用户不存在或验证码错误', code=400)

    cached_code = redis_client.get_forgot_password_code(user.id)
    if not cached_code or str(code).strip() != str(cached_code).strip():
        raise BusinessException('用户不存在或验证码错误', code=400)

    user.set_password(new_password)
    user.save()
    redis_client.delete_forgot_password_code(user.id)

    return Response({'code': 200, 'message': '密码重置成功'})
