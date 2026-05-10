"""
课程管理视图
"""
from rest_framework import viewsets
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import action
from rest_framework.parsers import MultiPartParser
from rest_framework.response import Response
from django.db import transaction
from django.db import models
import csv
import io
import json

from apps.users.models import User
from utils.exceptions import BusinessException
from utils.permissions import IsAdmin, IsTeacher
from utils.pagination import StandardResultsSetPagination
from .models import Course
from .serializers import CourseSerializer


class CourseViewSet(viewsets.ModelViewSet):
    """课程视图集"""
    queryset = Course.objects.all().order_by('-created_at')
    serializer_class = CourseSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = StandardResultsSetPagination

    def list(self, request, *args, **kwargs):
        """课程列表 - 返回标准响应格式"""
        queryset = self.filter_queryset(self.get_queryset())
        page = self.paginate_queryset(queryset)
        
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        
        serializer = self.get_serializer(queryset, many=True)
        return Response({
            'code': 200,
            'message': '获取成功',
            'data': serializer.data
        })
    
    def get_paginated_response(self, data):
        """自定义分页响应格式"""
        return Response({
            'code': 200,
            'message': '获取成功',
            'data': {
                'results': data,
                'count': self.paginator.page.paginator.count,
                'next': self.paginator.get_next_link(),
                'previous': self.paginator.get_previous_link(),
            }
        })

    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser])
    @transaction.atomic
    def bulk_import(self, request, *args, **kwargs):
        """批量导入课程（仅支持XLSX格式）"""
        
        rows_data = []
        
        # 检查是否是JSON数据（来自前端Excel解析）
        if 'data' in request.data:
            try:
                rows_data = json.loads(request.data.get('data'))
                if not isinstance(rows_data, list):
                    raise BusinessException('数据格式错误', code=400)
            except json.JSONDecodeError:
                raise BusinessException('JSON数据解析失败', code=400)
        
        # 检查是否是文件上传
        elif 'file' in request.data or len(request.FILES) > 0:
            # 优先从request.FILES获取文件
            if len(request.FILES) > 0:
                file_obj = request.FILES.get('file')
                if not file_obj:
                    # 尝试从request.data获取
                    file_obj = request.data.get('file')
            else:
                file_obj = request.data.get('file')
                
            if not file_obj:
                raise BusinessException('未上传文件', code=400)

            # 处理字符串和文件对象两种情况
            if hasattr(file_obj, 'name'):
                file_name = file_obj.name.lower()
            else:
                file_name = str(file_obj).lower()
            if not (file_name.endswith('.xlsx') or file_name.endswith('.xls')):
                raise BusinessException('只支持Excel文件格式（.xlsx/.xls）', code=400)

            # 处理Excel文件
            import openpyxl
            wb = openpyxl.load_workbook(file_obj)
            ws = wb.active
            
            # 获取表头
            header = [cell.value for cell in ws[1]]
            required_headers = ['code', 'name']
            if not all(h in header for h in required_headers):
                raise BusinessException(f"Excel文件头必须包含: {', '.join(required_headers)}", code=400)
            
            # 读取数据
            data_rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                data_rows.append(row)
            
            # 转换为字典格式
            rows_data = [dict(zip(header, row)) for row in data_rows]
        else:
            raise BusinessException('请上传Excel文件或提供数据', code=400)

        results = {
            'success_count': 0,
            'failure_count': 0,
            'errors': []
        }

        teacher_cache = {u.username: u for u in User.objects.filter(role=1)}

        for i, row_data in enumerate(rows_data, start=2):
            code = (row_data.get('code') or '').strip()
            name = (row_data.get('name') or '').strip()
            description = (row_data.get('description') or '').strip()
            teacher_username = (row_data.get('teacher_username') or '').strip()

            if not code or not name:
                results['failure_count'] += 1
                results['errors'].append({'row': i, 'code': code, 'message': '课程代码和课程名称不能为空'})
                continue

            if Course.objects.filter(code=code).exists():
                results['failure_count'] += 1
                results['errors'].append({'row': i, 'code': code, 'message': '课程代码已存在'})
                continue

            teacher = None
            if teacher_username:
                teacher = teacher_cache.get(teacher_username)
                if not teacher:
                    results['failure_count'] += 1
                    results['errors'].append({'row': i, 'code': code, 'message': f'教师账号 "{teacher_username}" 不存在或不是教师'})
                    continue

            try:
                Course.objects.create(
                    code=code,
                    name=name,
                    description=description or None,
                    teacher=teacher
                )
                results['success_count'] += 1
            except Exception as e:
                results['failure_count'] += 1
                results['errors'].append({'row': i, 'code': code, 'message': f'创建失败: {str(e)}'})

        return Response({
            'code': 200,
            'message': '导入完成',
            'data': results
        })

    @action(detail=False, methods=['delete'])
    @transaction.atomic
    def bulk_delete(self, request, *args, **kwargs):
        """
        批量删除课程
        """
        ids = request.data.get('ids', [])
        
        if not ids:
            return Response({'code': 400, 'message': '请提供要删除的课程ID列表'}, 
                           status=400)
        
        if not isinstance(ids, list):
            return Response({'code': 400, 'message': 'ids参数必须是数组'}, 
                           status=400)
        
        # 只有管理员可以批量删除
        if not request.user.is_staff and request.user.role != 0:
            return Response({'code': 403, 'message': '只有管理员可以批量删除课程'}, 
                           status=403)
        
        success_count = 0
        failed_count = 0
        failed_ids = []
        errors = []
        
        for course_id in ids:
            try:
                course = Course.objects.get(id=course_id)
                
                # 检查关联数据
                from apps.papers.models import Paper
                from apps.classes.models import Class
                
                # 检查是否有试卷使用该课程
                paper_count = Paper.objects.filter(course=course).count()
                if paper_count > 0:
                    errors.append(f'课程ID {course_id}: 该课程被 {paper_count} 份试卷使用，无法删除')
                    failed_count += 1
                    failed_ids.append(course_id)
                    continue
                
                # 检查是否有班级使用该课程
                class_count = Class.objects.filter(courses=course).count()
                if class_count > 0:
                    errors.append(f'课程ID {course_id}: 该课程被 {class_count} 个班级使用，无法删除')
                    failed_count += 1
                    failed_ids.append(course_id)
                    continue
                
                # 删除课程
                course.delete()
                success_count += 1
                
            except Course.DoesNotExist:
                errors.append(f'课程ID {course_id}: 课程不存在')
                failed_count += 1
                failed_ids.append(course_id)
            except Exception as e:
                errors.append(f'课程ID {course_id}: 删除失败 - {str(e)}')
                failed_count += 1
                failed_ids.append(course_id)
        
        return Response({
            'code': 200,
            'message': f'批量删除完成，成功 {success_count} 个，失败 {failed_count} 个',
            'data': {
                'success_count': success_count,
                'failed_count': failed_count,
                'failed_ids': failed_ids,
                'errors': errors
            }
        })
    
    def get_permissions(self):
        """
        根据操作类型设置不同的权限
        - list, retrieve: 管理员和教师都可以访问
        - create, update, destroy: 只有管理员可以访问
        """
        if self.action in ['list', 'retrieve']:
            # 列表和详情：管理员和教师都可以访问
            return [IsAuthenticated()]
        else:
            # 创建、更新、删除：只有管理员可以
            return [IsAuthenticated(), IsAdmin()]
    
    def get_queryset(self):
        """根据用户角色和筛选条件过滤课程"""
        queryset = super().get_queryset()
        
        # 如果是教师，只能看到自己教授的课程
        if self.request.user.role == 1:
            queryset = queryset.filter(teacher=self.request.user)
        # 如果是管理员，可以看到所有课程（不需要过滤）
        # 如果是学生，不应该访问这个接口（但权限已经限制了）
        
        # 搜索条件
        search = self.request.query_params.get('search')
        if search:
            queryset = queryset.filter(
                models.Q(name__icontains=search) |
                models.Q(code__icontains=search)
            )
        
        # 教师筛选
        teacher_id = self.request.query_params.get('teacher_id')
        if teacher_id:
            queryset = queryset.filter(teacher_id=teacher_id)
        
        return queryset